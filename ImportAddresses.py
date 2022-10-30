import sys
from subprocess import check_call, check_output
import os
import csv
import shutil
import time
import concurrent.futures
import threading


t1 = time.perf_counter()

def install_openpyxl():
    print("\u001b[33m\nInstalling openpyxl...")
    # implement pip as a subprocess:
    check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])

    # process output with an API in the subprocess module:
    reqs = check_output([sys.executable, '-m', 'pip','freeze'])
    installed_packages = [r.decode().split('==')[0] for r in reqs.split()]

    print(installed_packages)

try:
    import openpyxl
except ModuleNotFoundError:
    print("\u001b[31;1mOpenpyxl library is not installed.")
    install_openpyxl()


import openpyxl

#define the progress bar class
class progressBar:

    prog = 0
    total = 0
        
    def print_progress_bar():
        while progressBar.prog < progressBar.total:
            percent = round((progressBar.prog / progressBar.total) * 100)
            bar = '█' * int(percent) + '-' * (100 - int(percent))#'█'
            print(f"\r|{bar}| {percent:.0f}%", end="\r", flush=True)
            if percent == 100:
                break

    def __init__(self):
        pass


#print title and check python version
def preamble():
    os.system('color')
    print("\u001b[4m\u001b[35;1mEvents Layout Import Script\u001b[0m")
    #print("\u001b[37m\u001b[0mPython Version " + sys.version)
    if sys.version[:7] != "3.10.8 ":
        print("\u001b[33;1m***Warning: The version of Python is different from what this script was written on.***")


#Confirming, finding, and copying files.
def manages_files():
    wrk_dir = os.getcwd()
    temp_dir = wrk_dir + '//template'
    out_dir = wrk_dir + '//output'
    in_dir = wrk_dir + '//input'
    #confirming files
    try:
        temp_loc = temp_dir + '//' + os.listdir(temp_dir)[0]
    except FileNotFoundError:
        print("\n")
        print("\u001b[1m\u001b[31;1mThe template file or directory was not found.\n\nPlease add the template file to the template directory and restart.")
        done()
    except IndexError:
        print("\n")
        print("\u001b[1m\u001b[31;1mThe template file was not found.\n\nPlease add the template file to the template directory and restart.")
        done()
    try:
        in_loc = in_dir + '//' + os.listdir(in_dir)[0]
    except FileNotFoundError:
        print("\n")
        print("\u001b[1m\u001b[31;1mThe input file or directory was not found.\n\nPlease add the input file to the input directory and restart.")
        done()
    except IndexError:
        print("\n")
        print("\u001b[1m\u001b[31;1mThe input file was not found.\n\nPlease add the input file to the input directory and restart.")
        done()
    #Copying template file to output directory
    try:
        shutil.copy(temp_loc, out_dir + '//out_' + os.listdir(temp_dir)[0])
    except FileNotFoundError:
        
        print("\n")
        print("\u001b[1m\u001b[31;1mThe output directory was not found.\n\nPlease add the output directory and restart.")
        done()
    out_loc = out_dir + '//' + os.listdir(out_dir)[0]
    locations = [temp_loc, out_loc, in_loc]
    return locations


#Resets the text color
def done():
    print("\u001b[37m\u001b[0m")
    t2 = time.perf_counter()
    time_elapsed = round((t2 - t1), 3)
    print("Execution time: " + f"{time_elapsed}" + "sec(s)")
    #input("Press Enter to close window...")
    exit()


#gets address that need comments
def get_address_array_from_temp(sheet):
    array = []
    for i in range(sheet.max_row):
        addy = sheet.cell(i+3,2).value
        if type(addy) != str:
            continue
        elif addy[0:3] == "GMF":
            array.append([addy, i + 3])# [0]requested address [1]position
        elif addy[0:2] == "EM":
            array.append([addy, i + 3])#adding 3 to index number to offset for formating of template
        else:
            pass
    return array


#gets all address that have comments
def get_address_comment_array_from_input(location):
    try:
        array = list(csv.reader(open(location, encoding= "ISO8859")))
    except PermissionError:
        print("\u001b[1m\u001b[31mError: Could not access input file.")
        done()
    return array

#main code
def main():
    preamble()#run preamble

    file_locs = manages_files()#file_locs[template, output, input]

    #open output workbook and worksheet
    wb = openpyxl.load_workbook(filename=file_locs[1])
    ws = wb["Import Cheat Sheet"]

    #get address that need comments from template and get addresses with comments from input in seperate threads
    with concurrent.futures.ThreadPoolExecutor() as executor:
        f1 = executor.submit(get_address_array_from_temp, ws)
        f2 = executor.submit(get_address_comment_array_from_input, file_locs[2])
    #wait for results from both threads
    address_array = f1.result()
    address_comment_array = f2.result()
    #close executor
    executor.shutdown()

    match_count = 0
    address_array_len = len(address_array)
    
    print("\n\u001b[0m\u001b[32mWorking on it...",flush=True)

    #set the progress bar total and start the progress bar thread
    progressBar.total = address_array_len
    t1 = threading.Thread(target=progressBar.print_progress_bar)
    t1.start()

    #loop through the addresses and compare to the array with comments
    for i in range(address_array_len):
        for address in address_comment_array:
            if address_array[i][0] == address[0]:
                ws.cell(row=address_array[i][1], column=6).value = address[0]
                ws.cell(row=address_array[i][1], column=7).value = address[1]
                match_count+=1
            else:
                progressBar.prog = i
    
    #set progress on progress bar to 100
    progressBar.prog = address_array_len - 1
    
    #wait for progress bar thread to finish
    t1.join()
    
    #print the progress bar at 100%
    progressBar.print_progress_bar()
    
    #save changes to the ouput file
    wb.save(file_locs[1])
    
    #display stats and warning if needed
    print("\nDone.", flush=True)
    print("\n\u001b[34;1mNumber of comments found:\u001b[33m" + str(match_count))
    if match_count == 0:
        print("\u001b[33;1m***No matches were found. Make sure your input and template files are correct***")

    #reset and exit()
    done()
    #end of main


main()
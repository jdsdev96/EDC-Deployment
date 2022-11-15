#FILENAME:ImportEvents.py
#AUTHOR:Jonathan Shambaugh
#PURPOSE: To extract the comments given in a Toyopuc project and write them to the corresponding address in the template for easy event importing.
#NOTES: See the github repository for more info. https://github.com/jdsdev96/EDC-ImportEventsTool
#VERSION: v1.2.3
#START DATE: 17 Oct 22

from sys import executable, version
from subprocess import check_call, check_output
from os import system, getcwd, listdir, access, R_OK, W_OK, X_OK, stat
from csv import reader
from shutil import copy
from time import perf_counter
from concurrent.futures import ThreadPoolExecutor
from threading import Thread
from msvcrt import getch, kbhit



v = "v1.2.3"


t1 = perf_counter()



#installs openpyxl using the command line
def install_lib(lib):
    print(f"\u001b[33m\nInstalling {lib}...")
    # implement pip as a subprocess:
    check_call([executable, '-m', 'pip', 'install', lib])

    # process output with an API in the subprocess module:
    reqs = check_output([executable, '-m', 'pip','freeze'])
    installed_packages = [r.decode().split('==')[0] for r in reqs.split()]

    print(installed_packages)

#check if openpyxl is installed, if not, install it
try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    print("\u001b[31;1mOpenpyxl library is not installed.")
    install_lib("Openpyxl")
    from openpyxl import load_workbook
try:
    from requests import get
except ModuleNotFoundError:
    print("\u001b[31;1mRequests library is not installed.")
    install_lib("requests")
    from requests import get


#define the progress bar class
class progressBar:

    def print_progress_bar(self):
        while self.prog < self.total:
            percent = round((self.prog / self.total) * 100)
            bar = '█' * int(percent) + '-' * (100 - int(percent))#'█'
            print(f"\r|{bar}| {percent:.0f}%", end="\r", flush=True)
            if percent == 100:
                break
        return None

    def __init__(self, prog, total):
        self.prog = prog
        self.total = total



#print title and check python version
def preamble():
    system('color')
    print("\u001b[4m\u001b[35;1mEvents Layout Import Tool\u001b[0m")
    print(v)
    #print("\u001b[37m\u001b[0mPython Version: " + version[:7])
    if version[:4] != "3.10":
        print("\u001b[33;1m***Warning: The version of Python is different from what this script was written on.***")
        return None
    owner = "jdsdev96"
    repo = "EDC-ImportEventsTool"
    print("Checking for updates...", end="",flush=True)
    try:
        response = get(f"https://api.github.com/repos/{owner}/{repo}/releases/latest")
        #print(response.json())
        print("[DONE]")
        if v != response.json()["tag_name"]:
            print("\u001b[33;1m***Warning: There is a new release of this tool.***")
    except:
        print("[FAILED]")
        print("\u001b[33;1m***Warning: Could not connect to repository. Version check failed.***")
    #print(environ)


def get_current_cursor_pos():
    print("\n")
    print("\033[A\033[6n")
    keep_going = True
    buff = ""
    while keep_going:
        buff += getch().decode("ASCII")
        keep_going = kbhit()
    newbuff =buff.replace("\x1b[", "")
    return [newbuff[0], newbuff[2]]


#Confirming, finding, and copying files.
def manages_files():
    wrk_dir = getcwd()
    temp_dir, out_dir, in_dir = wrk_dir + '//template', wrk_dir + '//output', wrk_dir + '//input'
    #confirming files
    try:
        temp_loc = temp_dir + '//' + listdir(temp_dir)[0]
    except FileNotFoundError:
        print("\n")
        print("\u001b[1m\u001b[31;1mThe template directory was not found.\n\nPlease add the template directory and restart.")
        done()
    except IndexError:
        print("\n")
        print("\u001b[1m\u001b[31;1mThe template file was not found.\n\nPlease add the template file to the template directory and restart.")
        done()
    try:
        in_loc = in_dir + '//' + listdir(in_dir)[0]
    except FileNotFoundError:
        print("\n")
        print("\u001b[1m\u001b[31;1mThe input file or directory was not found.\n\nPlease add the input file to the input directory and restart.")
        done()
    except IndexError:
        print("\n")
        print("\u001b[1m\u001b[31;1mThe input file was not found.\n\nPlease add the input file to the input directory and restart.")
        done()
    #Copying template file to output directory
    try:
        copy(temp_loc, out_dir + '//out_' + listdir(temp_dir)[0])
    except FileNotFoundError:
        print("\n")
        print("\u001b[1m\u001b[31;1mThe output directory was not found.\n\nPlease add the output directory and restart.")
        done()
    except Exception as e:
        print("Make sure to close the template file or make sure template file is not being used by another program.")
        print(e)
        done()


    out_loc = out_dir + '//' + listdir(out_dir)[0]
    locations = [temp_loc, out_loc, in_loc]
    return locations


def perm_check(locs):
    file_names = ["template", "output", "input"]
    access_type = ["read", "write", "execute"]
    for i in range(len(locs)):
        permissions = [access(locs[i], R_OK), access(locs[i], W_OK), access(locs[i], X_OK)]
        for j in range(len(permissions)):
            if not permissions[j]:
                print(f"\u001b[1m\u001b[31;1mThe script does not have {access_type[j]} access to the {file_names[i]} file. Make sure the file is closed and permissions are set.")
            else:
                #print(f"\u001b[1m\u001b[31;1mThe script does have {access_type[j]} access to the {file_names[i]} file. Make sure the file is closed and permissions are set.")
                continue
        continue
    return None


#gets address that need comments
def get_address_array_from_temp(sheet):
    array = []
    for i in range(sheet.max_row):
        array.append([sheet.cell(i+3,2).value, i + 3])
    return array


#gets all address that have comments
def get_address_comment_array_from_input(location):
    try:
        array = list(reader(open(location, encoding= "ISO8859")))
    except PermissionError:
        print("\u001b[1m\u001b[31mError: Could not access input file.")
        done()
    return array


#Resets the text color
def done():
    print("\u001b[37m\u001b[0m")
    time_elapsed = round((perf_counter() - t1), 3)
    print(" ".join(["Execution time: ", f"{time_elapsed}", "sec(s)"]))
    #input("throwaway")
    exit()


#main code
def main():
    #run preamble
    preamble()

    #find file locations
    file_locs = manages_files()#file_locs[template, output, input]

    #check permissions on files
    perm_check(file_locs)

    #open output workbook and worksheet
    wb = load_workbook(filename=file_locs[1])
    ws = wb["Import Cheat Sheet"]

    #get address that need comments from template and get addresses with comments from input in separate threads
    with ThreadPoolExecutor() as executor:
        f1 = executor.submit(get_address_array_from_temp, ws)
        f2 = executor.submit(get_address_comment_array_from_input, file_locs[2])
    #wait for results from both threads
    address_array = f1.result()
    address_comment_array = f2.result()
    #close executor
    executor.shutdown()

    match_count, address_array_len = 0, len(address_array)
    
    print("\n\u001b[0m\u001b[32mWorking on it...",flush=True)

    #set the progress bar total and start the progress bar thread
    address_prog_bar = progressBar(0, address_array_len)
    t1 = Thread(target=address_prog_bar.print_progress_bar)
    t1.start()

    #loop through the addresses and compare to the array with comments
    for i in range(address_array_len):
        for address in address_comment_array:
            if address_array[i][0] == address[0]:
                ws.cell(row=address_array[i][1], column=6).value = address[0]
                ws.cell(row=address_array[i][1], column=7).value = address[1]
                match_count+=1
            elif address[0][:4] == "P1-X":
                if address_array[i][0] == address[0][3:]:
                    ws.cell(row=address_array[i][1], column=6).value = address[0][3:]
                    ws.cell(row=address_array[i][1], column=7).value = address[1]
                    match_count+=1
                else:
                    continue
            elif address[0][:4] == "P2-D":
                if address_array[i][0] == address[0][3:]:
                    ws.cell(row=address_array[i][1], column=6).value = address[0][3:]
                    ws.cell(row=address_array[i][1], column=7).value = address[1]
                    match_count+=1
                else:
                    continue
            else:
                address_prog_bar.prog = i
    
    #set progress on progress bar to 100
    address_prog_bar.prog = address_array_len - 1
    
    #wait for progress bar thread to finish
    t1.join()
    
    #print the progress bar at 100%
    address_prog_bar.print_progress_bar()
    
    #save changes to the output file
    wb.save(file_locs[1])
    
    #display stats and warning if needed
    print("\nDone.", flush=True)
    print("\n\u001b[34;1mNumber of comments found:\u001b[33m" + str(match_count))
    if match_count == 0:
        print("\u001b[33;1m***No matches were found. Make sure your input and template files are correct***")

    #reset and exit()
    done()
    #end of main


main()